import pandas as pd
from openpyxl import load_workbook
from IPython.display import Markdown, display

excel_file = 'DB_Objects.xlsx'

df1 = pd.read_excel(excel_file, sheet_name='AdventureWorks2014')
df2 = pd.read_excel(excel_file, sheet_name='AdventureWorksDW2014')

# Step 1: Combine both lists of SQL objects (including OBJ_TYPE and OBJ_NAME)
# Create a union of all unique (OBJ_TYPE, OBJ_NAME) pairs
df_all_objects = pd.merge(df1[['OBJ_TYPE', 'OBJ_NAME']], df2[['OBJ_TYPE', 'OBJ_NAME']], 
                       how='outer', on=['OBJ_TYPE', 'OBJ_NAME'], indicator=True)

# Step 2: Create the result DataFrame with columns for "SERVER01" and "SERVER02"
df_all_objects['AdventureWorks2014'] = df_all_objects['_merge'].apply(lambda x: 'YES' if x != 'right_only' else 'NO')
df_all_objects['AdventureWorksDW2014'] = df_all_objects['_merge'].apply(lambda x: 'YES' if x != 'left_only' else 'NO')

# Drop the '_merge' column as it's no longer needed
df_all_objects.drop(columns=['_merge'], inplace=True)

# Step 3: Sort by OBJ_TYPE and OBJ_NAME
df_all_objects = df_all_objects.sort_values(by=['OBJ_TYPE', 'OBJ_NAME'])

# Step 4: Create new sheet in existing excel file to store the comparison

sheet_name_output = "Object Comparison"

try:
    book = load_workbook(excel_file)
    if sheet_name_output in book.sheetnames:
        del book[sheet_name_output]

    # Writing to a new sheet in the Excel file
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Write the DataFrame to a new sheet named 'NewSheet'
        df_all_objects.to_excel(writer, sheet_name=sheet_name_output, index=False)
        
except FileNotFoundError:
    print(f"Error: The file {excel_file} does not exist.")

# Print message of completion


def PrintFormatedText(string, color=None):
    colorstr = "<span style='color:{}'>{}</span>".format(color, string)
    display(Markdown(colorstr))
    
PrintFormatedText("**Process Completed Successfully!!!**", color="green")